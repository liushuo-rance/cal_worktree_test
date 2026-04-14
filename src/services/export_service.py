"""
导出服务
提供 CSV、Excel、PDF 导出功能
"""

import csv
import io
import os
import platform
import urllib.request
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


class ExportError(Exception):
    """导出服务异常"""
    pass


# 中文字体查找路径（按优先级排列）
_CHINESE_FONT_PATHS = {
    "Darwin": [
        "/System/Library/Fonts/PingFang.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
    ],
    "Linux": [
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ],
    "Windows": [
        r"C:\\Windows\\Fonts\\msyh.ttc",
        r"C:\\Windows\\Fonts\\simsun.ttc",
        r"C:\\Windows\\Fonts\\simhei.ttf",
    ],
}

_NOTO_FONT_URL = (
    "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/SimplifiedChinese/"
    "NotoSansCJKsc-Regular.otf"
)


def _ensure_noto_font(static_fonts_dir: str) -> str:
    """下载 Noto Sans CJK 字体到指定目录并返回路径"""
    os.makedirs(static_fonts_dir, exist_ok=True)
    font_path = os.path.join(static_fonts_dir, "NotoSansCJKsc-Regular.otf")
    if os.path.exists(font_path):
        return font_path
    try:
        urllib.request.urlretrieve(_NOTO_FONT_URL, font_path)
    except Exception as exc:
        raise ExportError(f"下载中文字体失败: {exc}") from exc
    return font_path


def get_chinese_font() -> str:
    """
    按平台查找可用的中文字体路径。

    优先搜索系统自带字体，未找到时自动下载 Noto Sans CJK 到
    src/web/static/fonts/ 目录。

    Returns:
        可用字体文件的绝对路径

    Raises:
        ExportError: 完全无法获取可用字体时抛出
    """
    system = platform.system()
    candidates = _CHINESE_FONT_PATHS.get(system, [])

    for path in candidates:
        # reportlab 的 TTFont 对 .ttc 支持不稳定，优先使用 .ttf/.otf
        if os.path.isfile(path) and not path.lower().endswith(".ttc"):
            return path

    # 回退到项目目录中的静态字体
    fallback_dir = os.path.join(
        os.path.dirname(__file__), "..", "web", "static", "fonts"
    )
    fallback_dir = os.path.abspath(fallback_dir)
    font_path = os.path.join(fallback_dir, "NotoSansCJKsc-Regular.otf")

    if os.path.exists(font_path):
        return font_path

    # 尝试下载
    try:
        return _ensure_noto_font(fallback_dir)
    except ExportError:
        pass

    # 如果仍未找到，抛出异常
    raise ExportError(
        "未能找到可用的中文字体，请手动安装中文字体或将字体文件放置到 "
        f"{fallback_dir} 目录下"
    )


def export_to_csv(
    data: List[Dict[str, Any]],
    columns: List[Tuple[str, str]],
    output_path: Optional[str] = None,
) -> Optional[bytes]:
    """
    将数据导出为 CSV 格式（UTF-8 with BOM）。

    Args:
        data: 待导出的记录列表，每条记录为字典
        columns: [(db_key, chinese_header), ...]，控制列顺序与中文表头
        output_path: 若提供，则写入该文件路径；否则返回 bytes

    Returns:
        若 output_path 为 None，返回 CSV 文件内容的 bytes；否则返回 None
    """
    if not columns:
        raise ExportError("columns 不能为空")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([header for _, header in columns])

    for row in data:
        writer.writerow([row.get(key, "") for key, _ in columns])

    content = output.getvalue()
    # UTF-8 with BOM，确保 Excel 直接打开中文不乱码
    encoded = content.encode("utf-8-sig")

    if output_path:
        with open(output_path, "wb") as f:
            f.write(encoded)
        return None

    return encoded


def export_to_excel(
    sheets_data: Dict[str, Dict[str, Any]],
    output_path: Optional[str] = None,
) -> Optional[bytes]:
    """
    将多 sheet 数据导出为 Excel（.xlsx）。

    Args:
        sheets_data: {
            'sheet名称': {
                'columns': [(db_key, chinese_header), ...],
                'data': [dict, ...]
            },
            ...
        }
        output_path: 若提供，则写入该文件路径；否则返回 bytes

    Returns:
        若 output_path 为 None，返回 Excel 文件内容的 bytes；否则返回 None
    """
    wb = Workbook()
    first = True

    for sheet_name, sheet_info in sheets_data.items():
        columns: List[Tuple[str, str]] = sheet_info.get("columns", [])
        data: List[Dict[str, Any]] = sheet_info.get("data", [])

        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 写入表头
        ws.append([header for _, header in columns])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for row in data:
            ws.append([row.get(key, "") for key, _ in columns])

        # 自动调整列宽
        for idx, (key, header) in enumerate(columns, start=1):
            col_letter = get_column_letter(idx)
            max_length = len(str(header))
            for row in data:
                val_len = len(str(row.get(key, "")))
                if val_len > max_length:
                    max_length = val_len
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    if output_path:
        with open(output_path, "wb") as f:
            f.write(excel_bytes)
        return None

    return excel_bytes


def _format_currency(value: float) -> str:
    """格式化为两位小数的人民币金额"""
    return f"¥{value:.2f}"


def export_report_to_pdf(
    report_data: Dict[str, Any],
    report_type: str,
    output_path: Optional[str] = None,
) -> Optional[bytes]:
    """
    将报表数据导出为 PDF。

    Args:
        report_data: 来自 report_service.py 对应报表函数的结果字典
        report_type: 'monthly' | 'salary' | 'comp_off'
        output_path: 若提供，则写入该文件路径；否则返回 bytes

    Returns:
        若 output_path 为 None，返回 PDF 文件内容的 bytes；否则返回 None
    """
    if report_type == "monthly":
        return _render_monthly_pdf(report_data, output_path)
    if report_type == "salary":
        return _render_salary_pdf(report_data, output_path)
    if report_type == "comp_off":
        return _render_comp_off_pdf(report_data, output_path)

    raise ExportError(f"不支持的报表类型: {report_type}")


def _render_monthly_pdf(
    report_data: Dict[str, Any],
    output_path: Optional[str] = None,
) -> Optional[bytes]:
    """生成月度加班报表 PDF"""
    employee_id = report_data.get("employee_id", "")
    employee_name = report_data.get("employee_name", "")
    year = report_data.get("year", "")
    month = report_data.get("month", "")
    summary = report_data.get("summary", {})
    overtime_details = report_data.get("overtime_details", [])
    leave_details = report_data.get("leave_details", [])

    elements = []
    styles = getSampleStyleSheet()
    font_name = _register_chinese_font()

    title_style = styles["Title"]
    title_style.fontName = font_name
    title_style.fontSize = 18

    normal_style = styles["Normal"]
    normal_style.fontName = font_name
    normal_style.fontSize = 10

    # 标题
    elements.append(Paragraph("月度加班报表", title_style))
    elements.append(Spacer(1, 0.3 * cm))

    # 元信息
    meta_text = (
        f"员工: {employee_name} ({employee_id}) &nbsp;&nbsp; "
        f"报表期间: {year}年{month}月 &nbsp;&nbsp; "
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    elements.append(Paragraph(meta_text, normal_style))
    elements.append(Spacer(1, 0.5 * cm))

    # 汇总表格
    summary_data = [
        ["工作日加班(小时)", "周末加班(小时)", "节假日加班(小时)", "总加班(小时)", "请假(天)", "请假(小时)"],
        [
            summary.get("weekday_hours", 0),
            summary.get("weekend_hours", 0),
            summary.get("holiday_hours", 0),
            summary.get("total_overtime_hours", 0),
            summary.get("leave_days", 0),
            summary.get("leave_hours", 0),
        ],
    ]
    summary_table = Table(summary_data, colWidths=[2.5 * cm] * 6)
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("FONTNAME", (0, 1), (-1, -1), font_name),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 0.8 * cm))

    # 加班明细
    if overtime_details:
        elements.append(Paragraph("加班明细", normal_style))
        elements.append(Spacer(1, 0.2 * cm))
        ot_data = [["日期", "星期", "时长(小时)", "类型", "说明"]]
        for row in overtime_details:
            ot_data.append(
                [
                    row.get("date", ""),
                    row.get("weekday", ""),
                    row.get("hours", 0),
                    row.get("type", ""),
                    row.get("description", ""),
                ]
            )
        ot_table = Table(ot_data, colWidths=[3 * cm, 2 * cm, 2.5 * cm, 3 * cm, 5 * cm])
        ot_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("ALIGN", (-2, 1), (-2, -1), "LEFT"),
                    ("ALIGN", (-1, 1), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), font_name),
                    ("FONTNAME", (0, 1), (-1, -1), font_name),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        elements.append(ot_table)
        elements.append(Spacer(1, 0.5 * cm))

    # 请假明细
    if leave_details:
        elements.append(Paragraph("请假明细", normal_style))
        elements.append(Spacer(1, 0.2 * cm))
        leave_data = [["日期", "星期", "时长(小时)", "天数", "类型"]]
        for row in leave_details:
            leave_data.append(
                [
                    row.get("date", ""),
                    row.get("weekday", ""),
                    row.get("hours", 0),
                    row.get("days", 0),
                    row.get("type", ""),
                ]
            )
        leave_table = Table(leave_data, colWidths=[3 * cm, 2 * cm, 2.5 * cm, 2 * cm, 5.5 * cm])
        leave_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("ALIGN", (-1, 1), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), font_name),
                    ("FONTNAME", (0, 1), (-1, -1), font_name),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        elements.append(leave_table)

    return _build_pdf(elements, output_path)


def _render_salary_pdf(
    report_data: Dict[str, Any],
    output_path: Optional[str] = None,
) -> Optional[bytes]:
    """生成工资计算表 PDF"""
    employee_id = report_data.get("employee_id", "")
    employee_name = report_data.get("employee_name", "")
    year = report_data.get("year", "")
    month = report_data.get("month", "")
    hourly_rate = report_data.get("hourly_rate", 0)

    elements = []
    styles = getSampleStyleSheet()
    font_name = _register_chinese_font()

    title_style = styles["Title"]
    title_style.fontName = font_name
    title_style.fontSize = 18

    normal_style = styles["Normal"]
    normal_style.fontName = font_name
    normal_style.fontSize = 10

    elements.append(Paragraph("工资计算表", title_style))
    elements.append(Spacer(1, 0.3 * cm))

    meta_text = (
        f"员工: {employee_name} ({employee_id}) &nbsp;&nbsp; "
        f"报表期间: {year}年{month}月 &nbsp;&nbsp; "
        f"时薪: {_format_currency(hourly_rate)} &nbsp;&nbsp; "
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    elements.append(Paragraph(meta_text, normal_style))
    elements.append(Spacer(1, 0.5 * cm))

    salary_data = [
        ["项目", "加班时长(小时)", "倍数", "金额"],
        [
            "工作日加班",
            report_data["weekday_overtime"]["hours"],
            report_data["weekday_overtime"]["multiplier"],
            _format_currency(report_data["weekday_overtime"]["amount"]),
        ],
        [
            "周末加班",
            report_data["weekend_overtime"]["hours"],
            report_data["weekend_overtime"]["multiplier"],
            _format_currency(report_data["weekend_overtime"]["amount"]),
        ],
        [
            "节假日加班",
            report_data["holiday_overtime"]["hours"],
            report_data["holiday_overtime"]["multiplier"],
            _format_currency(report_data["holiday_overtime"]["amount"]),
        ],
        ["合计", "", "", _format_currency(report_data["total_amount"])],
    ]

    salary_table = Table(salary_data, colWidths=[4 * cm, 3.5 * cm, 2.5 * cm, 3.5 * cm])
    salary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), font_name),
                ("FONTNAME", (0, 1), (-1, -1), font_name),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, -1), (-1, -1), font_name),
                ("TEXTCOLOR", (-1, -1), (-1, -1), colors.red),
            ]
        )
    )
    elements.append(salary_table)
    return _build_pdf(elements, output_path)


def _render_comp_off_pdf(
    report_data: Dict[str, Any],
    output_path: Optional[str] = None,
) -> Optional[bytes]:
    """生成调休余额报表 PDF"""
    employee_id = report_data.get("employee_id", "")
    employee_name = report_data.get("employee_name", "")
    summary = report_data.get("summary", {})
    warning_days = report_data.get("warning_days", 30)
    expiring_soon_hours = report_data.get("expiring_soon_hours", 0)
    active_balances = report_data.get("active_balances", [])

    elements = []
    styles = getSampleStyleSheet()
    font_name = _register_chinese_font()

    title_style = styles["Title"]
    title_style.fontName = font_name
    title_style.fontSize = 18

    normal_style = styles["Normal"]
    normal_style.fontName = font_name
    normal_style.fontSize = 10

    elements.append(Paragraph("调休余额报表", title_style))
    elements.append(Spacer(1, 0.3 * cm))

    meta_text = (
        f"员工: {employee_name} ({employee_id}) &nbsp;&nbsp; "
        f"累计获得: {summary.get('total_acquired_hours', 0)} 小时 &nbsp;&nbsp; "
        f"当前可用: {summary.get('total_available_hours', 0)} 小时 &nbsp;&nbsp; "
        f"{warning_days}天内到期: {expiring_soon_hours} 小时 &nbsp;&nbsp; "
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    elements.append(Paragraph(meta_text, normal_style))
    elements.append(Spacer(1, 0.5 * cm))

    if active_balances:
        comp_data = [["获得日期", "总时长(小时)", "已使用(小时)", "剩余(小时)", "到期日期", "状态"]]
        for row in active_balances:
            comp_data.append(
                [
                    row.get("acquired_date", ""),
                    row.get("total_hours", 0),
                    row.get("used_hours", 0),
                    row.get("remaining_hours", 0),
                    row.get("expiry_date", "") or "-",
                    row.get("status", ""),
                ]
            )
        col_widths = [3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 3 * cm, 2 * cm]
        comp_table = Table(comp_data, colWidths=col_widths)
        comp_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), font_name),
                    ("FONTNAME", (0, 1), (-1, -1), font_name),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]
            )
        )
        elements.append(comp_table)
    else:
        elements.append(Paragraph("暂无调休余额记录", normal_style))

    return _build_pdf(elements, output_path)


def _register_chinese_font() -> str:
    """注册中文字体到 reportlab 并返回字体名称"""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_path = get_chinese_font()
    # 使用固定的内部注册名称
    font_name = "ChineseFont"
    try:
        pdfmetrics.registerFont(TTFont(font_name, font_path))
    except Exception as exc:
        raise ExportError(f"注册中文字体失败: {exc}") from exc
    return font_name


def _build_pdf(elements: List[Any], output_path: Optional[str] = None) -> Optional[bytes]:
    """使用 reportlab 将元素列表构建为 PDF"""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    doc.build(elements)

    pdf_bytes = buf.getvalue()
    if output_path:
        with open(output_path, "wb") as f:
            f.write(pdf_bytes)
        return None
    return pdf_bytes
