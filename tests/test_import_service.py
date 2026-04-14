"""
导入服务测试
覆盖CSV、Excel、列映射、异常行、日期解析等场景
"""

import csv
import os
from datetime import date

import pytest

from src.services.import_service import (
    read_csv_file,
    read_excel_file,
    normalize_import_rows,
    _parse_date,
    _parse_hours,
    _parse_minutes,
    _build_alias_map,
    _infer_type,
    HAS_OPENPYXL,
)


class TestBuildAliasMap:
    def test_english_headers(self):
        headers = ["date", "employee_id", "hours", "type", "description"]
        result = _build_alias_map(headers)
        assert result["date"] == "date"
        assert result["employee_id"] == "employee_id"
        assert result["hours"] == "hours"
        assert result["type"] == "type"
        assert result["description"] == "description"

    def test_chinese_headers(self):
        headers = ["日期", "员工ID", "时长", "类型", "描述"]
        result = _build_alias_map(headers)
        assert result["date"] == "日期"
        assert result["employee_id"] == "员工ID"
        assert result["hours"] == "时长"
        assert result["type"] == "类型"
        assert result["description"] == "描述"

    def test_mixed_headers(self):
        headers = ["work_date", "工号", "duration_hours", "minutes", "备注"]
        result = _build_alias_map(headers)
        assert result["date"] == "work_date"
        assert result["employee_id"] == "工号"
        assert result["hours"] == "duration_hours"
        assert result["minutes"] == "minutes"
        assert result["description"] == "备注"

    def test_case_insensitive(self):
        headers = ["DATE", "HOURS", "TYPE"]
        result = _build_alias_map(headers)
        assert result["date"] == "DATE"
        assert result["hours"] == "HOURS"
        assert result["type"] == "TYPE"

    def test_partial_match(self):
        headers = ["时间", "小时"]
        result = _build_alias_map(headers)
        assert result["date"] == "时间"
        assert result["hours"] == "小时"
        assert "employee_id" not in result


class TestParseDate:
    def test_standard_format(self):
        assert _parse_date("2025-04-10") == "2025-04-10"

    def test_dot_separator(self):
        assert _parse_date("2025.04.10") == "2025-04-10"

    def test_slash_separator(self):
        assert _parse_date("2025/04/10") == "2025-04-10"

    def test_datetime_string(self):
        assert _parse_date("2025-04-10 14:30:00") == "2025-04-10"

    def test_date_object(self):
        assert _parse_date(date(2025, 4, 10)) == "2025-04-10"

    def test_empty_values(self):
        assert _parse_date(None) is None
        assert _parse_date("") is None
        assert _parse_date("   ") is None

    def test_invalid_date(self):
        assert _parse_date("not-a-date") is None


class TestParseHours:
    def test_integer(self):
        assert _parse_hours(2) == 2.0

    def test_float(self):
        assert _parse_hours(2.5) == 2.5

    def test_string_number(self):
        assert _parse_hours("3.5") == 3.5

    def test_chinese_half_day(self):
        assert _parse_hours("半天") == 4.0

    def test_chinese_full_day(self):
        assert _parse_hours("一天") == 8.0

    def test_chinese_hours(self):
        assert _parse_hours("2.5小时") == 2.5

    def test_chinese_hours_minutes(self):
        assert _parse_hours("2小时30分钟") == 2.5

    def test_empty(self):
        assert _parse_hours(None) is None
        assert _parse_hours("") is None

    def test_invalid(self):
        assert _parse_hours("abc") is None


class TestParseMinutes:
    def test_integer(self):
        assert _parse_minutes(30) == 30

    def test_float(self):
        assert _parse_minutes(30.4) == 30

    def test_string(self):
        assert _parse_minutes("45") == 45

    def test_empty(self):
        assert _parse_minutes(None) == 0
        assert _parse_minutes("") == 0

    def test_invalid(self):
        assert _parse_minutes("xyz") == 0


class TestReadCsvFile:
    def test_read_csv(self, tmp_path):
        file_path = tmp_path / "test.csv"
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["日期", "时长", "类型", "描述"])
            writer.writeheader()
            writer.writerow({"日期": "2025-04-10", "时长": "2.5", "类型": "加班", "描述": "项目文档"})
            writer.writerow({"日期": "2025-04-11", "时长": "3", "类型": "请假", "描述": "事假"})

        rows = read_csv_file(str(file_path))
        assert len(rows) == 2
        assert rows[0]["日期"] == "2025-04-10"
        assert rows[1]["时长"] == "3"

    def test_read_csv_semicolon(self, tmp_path):
        file_path = tmp_path / "test_semicolon.csv"
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write("date;hours;type\n")
            f.write("2025-04-10;2.5;overtime\n")

        rows = read_csv_file(str(file_path))
        assert len(rows) == 1
        assert rows[0]["date"] == "2025-04-10"


class TestReadExcelFile:
    @pytest.mark.skipif(not HAS_OPENPYXL, reason="缺少openpyxl")
    def test_read_xlsx(self, tmp_path):
        import openpyxl

        file_path = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["日期", "时长", "类型", "描述"])
        ws.append(["2025-04-10", 2.5, "加班", "项目文档"])
        ws.append(["2025-04-11", 3, "请假", "事假"])
        wb.save(str(file_path))

        rows = read_excel_file(str(file_path))
        assert len(rows) == 2
        assert rows[0]["日期"] == "2025-04-10"
        assert rows[1]["时长"] == 3

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="缺少openpyxl")
    def test_read_xlsx_with_empty_rows(self, tmp_path):
        import openpyxl

        file_path = tmp_path / "test_empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["date", "hours"])
        ws.append(["2025-04-10", 2])
        ws.append([None, None])
        ws.append(["2025-04-11", 3])
        wb.save(str(file_path))

        rows = read_excel_file(str(file_path))
        assert len(rows) == 2
        assert rows[0]["date"] == "2025-04-10"
        assert rows[1]["date"] == "2025-04-11"

    def test_read_excel_without_openpyxl(self, tmp_path):
        if HAS_OPENPYXL:
            pytest.skip("openpyxl已安装")
        file_path = tmp_path / "test.xlsx"
        with pytest.raises(ImportError):
            read_excel_file(str(file_path))


class TestNormalizeImportRows:
    def test_basic_overtime(self):
        rows = [
            {"日期": "2025-04-10", "时长": "2.5", "类型": "weekday_evening", "描述": "项目文档"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 1
        assert len(result["errors"]) == 0

        record = result["records"][0]
        assert record["type"] == "overtime"
        assert record["overtime_type"] == "weekday_evening"
        assert record["parsed_date"] == "2025-04-10"
        assert record["parsed_hours"] == 2.5
        assert record["total_minutes"] == 150
        assert record["content"] == "项目文档"
        assert record["confidence"] == 1.0
        assert record["confidence_level"] == "HIGH"
        assert record["is_valid"] is True

    def test_leave_record(self):
        rows = [
            {"date": "2025-04-11", "hours": "1", "type": "leave", "subtype": "sick", "description": "感冒"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 1
        record = result["records"][0]
        assert record["type"] == "leave"
        assert record["leave_type"] == "sick"
        assert record["parsed_hours"] == 1.0
        assert record["total_minutes"] == 60

    def test_comp_off_record(self):
        rows = [
            {"date": "2025-04-12", "hours": "4", "type": "comp_off", "description": "调休半天"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 1
        record = result["records"][0]
        assert record["type"] == "comp_off"
        assert record["parsed_hours"] == 4.0
        assert record["total_minutes"] == 240

    def test_hours_and_minutes(self):
        rows = [
            {"date": "2025-04-10", "hours": "2", "minutes": "30", "type": "overtime"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 1
        record = result["records"][0]
        assert record["parsed_hours"] == 2.5
        assert record["total_minutes"] == 150

    def test_chinese_type_inference(self):
        rows = [
            {"日期": "2025-04-10", "时长": "2", "类型": "周末", "描述": ""},
        ]
        result = normalize_import_rows(rows)
        record = result["records"][0]
        assert record["type"] == "overtime"
        assert record["overtime_type"] == "weekend"

    def test_employee_id_from_row_and_default(self):
        rows = [
            {"date": "2025-04-10", "hours": "2", "employee_id": "E001"},
        ]
        result = normalize_import_rows(rows)
        assert result["records"][0]["employee_id"] == "E001"

        rows2 = [{"date": "2025-04-10", "hours": "2"}]
        result2 = normalize_import_rows(rows2, employee_id="E002")
        assert result2["records"][0]["employee_id"] == "E002"

    def test_invalid_date(self):
        rows = [
            {"date": "invalid-date", "hours": "2"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 0
        assert len(result["errors"]) == 1
        assert "日期解析失败" in result["errors"][0]["message"]

    def test_missing_hours(self):
        rows = [
            {"date": "2025-04-10"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 0
        assert len(result["errors"]) == 1
        assert "无法解析时长" in result["errors"][0]["message"]

    def test_empty_rows_skipped(self):
        rows = [
            {"date": "2025-04-10", "hours": "2"},
            {"date": "", "hours": ""},
            {"date": "2025-04-11", "hours": "3"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 2

    def test_only_hours_column_with_zero_minutes(self):
        rows = [
            {"date": "2025-04-10", "hours": "0"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 0
        assert "无法解析时长" in result["errors"][0]["message"]

    def test_mixed_valid_and_invalid_rows(self):
        rows = [
            {"date": "2025-04-10", "hours": "2"},
            {"date": "bad-date", "hours": "3"},
            {"date": "2025-04-12", "hours": "4"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 2
        assert len(result["errors"]) == 1

    def test_subtype_from_type_column(self):
        rows = [
            {"date": "2025-04-10", "hours": "2", "type": "weekend"},
        ]
        result = normalize_import_rows(rows)
        record = result["records"][0]
        assert record["type"] == "overtime"
        assert record["overtime_type"] == "weekend"

    def test_normalization_with_no_subtype(self):
        rows = [
            {"date": "2025-04-10", "hours": "2"},
        ]
        result = normalize_import_rows(rows)
        record = result["records"][0]
        assert record["type"] == "overtime"
        assert record["overtime_type"] == "weekday_evening"

    def test_leaves_chinese_subtype(self):
        rows = [
            {"date": "2025-04-10", "hours": "8", "type": "leave", "subtype": "病假"},
        ]
        result = normalize_import_rows(rows)
        record = result["records"][0]
        assert record["type"] == "leave"
        assert record["leave_type"] == "sick"

    def test_overtime_chinese_subtype(self):
        rows = [
            {"date": "2025-04-10", "hours": "3", "type": "overtime", "subtype": "晚上"},
        ]
        result = normalize_import_rows(rows)
        record = result["records"][0]
        assert record["type"] == "overtime"
        assert record["overtime_type"] == "weekday_evening"

    def test_minutes_only(self):
        rows = [
            {"date": "2025-04-10", "minutes": "90"},
        ]
        result = normalize_import_rows(rows)
        assert len(result["records"]) == 1
        record = result["records"][0]
        assert record["total_minutes"] == 90
        assert record["parsed_hours"] == 1.5


class TestIntegration:
    def test_csv_to_normalized(self, tmp_path):
        file_path = tmp_path / "integration.csv"
        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["日期", "工号", "时长", "类型", "备注"])
            writer.writeheader()
            writer.writerow({"日期": "2025-04-10", "工号": "E001", "时长": "2.5", "类型": "weekend", "备注": "周末加班"})
            writer.writerow({"日期": "2025-04-11", "工号": "E001", "时长": "8", "类型": "leave", "备注": "事假"})

        raw_rows = read_csv_file(str(file_path))
        result = normalize_import_rows(raw_rows)

        assert len(result["records"]) == 2
        assert result["records"][0]["type"] == "overtime"
        assert result["records"][0]["employee_id"] == "E001"
        assert result["records"][1]["type"] == "leave"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="缺少openpyxl")
    def test_excel_to_normalized(self, tmp_path):
        import openpyxl

        file_path = tmp_path / "integration.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["date", "hours", "type", "description"])
        ws.append(["2025-04-10", 2.5, "overtime", "项目加班"])
        ws.append(["2025-04-11", 8, "annual", "年假"])
        wb.save(str(file_path))

        raw_rows = read_excel_file(str(file_path))
        result = normalize_import_rows(raw_rows)

        assert len(result["records"]) == 2
        assert result["records"][0]["type"] == "overtime"
        assert result["records"][1]["type"] == "leave"
        assert result["records"][1]["leave_type"] == "annual"
